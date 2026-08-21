"""
Professional Excel report generator.
Creates styled, color-coded recruitment reports with contact info,
skill analysis, and scoring breakdown.
"""

import io
from typing import Optional

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

from ..models.schemas import CandidateResult


# Color scheme
_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_GREEN_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
_AMBER_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
_RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
_GREEN_FONT = Font(name="Calibri", size=10, color="065F46")
_AMBER_FONT = Font(name="Calibri", size=10, color="92400E")
_RED_FONT = Font(name="Calibri", size=10, color="991B1B")


# Column definitions: (header, width, key_func)
_COLUMNS = [
    ("Rank", 6),
    ("Candidate Name", 22),
    ("Email", 28),
    ("Phone", 18),
    ("LinkedIn", 30),
    ("Location", 18),
    ("Final Score (%)", 14),
    ("Current Role", 24),
    ("Experience (Yrs)", 14),
    ("Verified Skills", 30),
    ("Listed-Only Skills", 24),
    ("Missing Skills", 24),
    ("Bonus Skills", 20),
    ("Skill Score (/35)", 14),
    ("Exp Score (/45)", 14),
    ("Keyword Score (/100)", 16),
    ("Summary", 50),
]


def _build_row(idx: int, result: CandidateResult) -> list:
    """Build a single data row from a CandidateResult."""
    audit = result.audit
    subscores = audit.subscores

    return [
        idx,
        result.candidate_name,
        result.contact.email,
        result.contact.phone,
        result.contact.linkedin,
        result.contact.location,
        result.final_score_pct,
        result.current_role,
        result.candidate_yoe,
        ", ".join(result.contextual_skills) or "None",
        ", ".join(result.stuffed_skills) or "None",
        ", ".join(result.missing_skills) or "None",
        ", ".join(result.nice_to_have_matched) or "None",
        subscores.skill_match,
        subscores.recent_exp,
        subscores.bm25_keyword,
        result.candidate_summary,
    ]


def generate_excel_report(
    results: list[CandidateResult],
) -> io.BytesIO:
    """
    Generate a professionally styled Excel report.

    Returns a BytesIO containing the .xlsx file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidate Rankings"

    # === Header Row ===
    for col_idx, (header, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # === Data Rows ===
    for idx, result in enumerate(results, start=1):
        row_data = _build_row(idx, result)
        row_num = idx + 1  # +1 for header

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(col_idx == len(_COLUMNS)),  # Wrap summary column
            )

        # Score-based row coloring
        score = result.final_score_pct
        if score >= 65:
            row_fill = _GREEN_FILL
        elif score >= 45:
            row_fill = _AMBER_FILL
        else:
            row_fill = _RED_FILL

        for col_idx in range(1, len(_COLUMNS) + 1):
            ws.cell(row=row_num, column=col_idx).fill = row_fill

        # Colored fonts for skill columns
        ws.cell(row=row_num, column=10).font = _GREEN_FONT   # Verified
        ws.cell(row=row_num, column=11).font = _AMBER_FONT   # Listed-only
        ws.cell(row=row_num, column=12).font = _RED_FONT      # Missing

    # === Alternating row shading (subtle) ===
    alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            score_cell = ws.cell(row=row_idx, column=7)
            try:
                score_val = float(score_cell.value)
                if 45 <= score_val < 65:
                    continue  # Keep amber coloring
            except (ValueError, TypeError):
                pass

    # === Add Analytics Chart ===
    if len(results) > 0:
        from openpyxl.chart import ScatterChart, Reference, Series
        ws_chart = wb.create_sheet(title="Analytics")
        chart = ScatterChart()
        chart.title = "Candidate Match vs Experience"
        chart.style = 13
        chart.x_axis.title = "Experience (Years)"
        chart.y_axis.title = "Final Score (%)"
        
        # Experience is col 9, Score is col 7
        xvalues = Reference(ws, min_col=9, min_row=2, max_row=len(results)+1)
        yvalues = Reference(ws, min_col=7, min_row=2, max_row=len(results)+1)
        
        series = Series(yvalues, xvalues, title_from_data=False)
        # Add labels to points if possible, though OpenPyXL scatter doesn't easily do data labels without complex xml
        chart.series.append(series)
        ws_chart.add_chart(chart, "B2")

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# Export init
__all__ = ["generate_excel_report"]
