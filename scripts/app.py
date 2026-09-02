import os
os.environ["USE_TF"] = "0"
os.environ["USE_TORCH"] = "1"

import io
import re
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import streamlit as st

# Core Engine Imports
from src.parser import ResumeParser, extract_must_haves_with_keybert, extract_skills_dual, extract_required_yoe
from src.scorer import HybridScorer

# 1. Page Configuration
st.set_page_config(
    page_title="AI ATS Resume Matcher", 
    page_icon="⚡", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- CENTRALIZED LAZY-LOADED MODEL REGISTRY ---
@st.cache_resource(show_spinner="Initializing AI Model Registry...")
def get_model_registry():
    from sentence_transformers import SentenceTransformer
    from keybert import KeyBERT

    transformer = SentenceTransformer("all-MiniLM-L6-v2")
    keybert_model = KeyBERT(model=transformer)
    scorer = HybridScorer(vector_model=transformer)
    parser = ResumeParser()
    return {
        "transformer": transformer,
        "keybert": keybert_model,
        "scorer": scorer,
        "parser": parser
    }

# --- TOP-LEVEL CACHED PARSING FUNCTION ---
@st.cache_data
def cached_parse_jd(jd_bytes: bytes, file_name: str):
    registry = get_model_registry()
    parser = registry["parser"]
    keybert_model = registry["keybert"]
    
    text = parser.parse_jd(jd_bytes, file_name=file_name)
    must_haves, nice_to_haves = extract_skills_dual(text, keybert_model=keybert_model)
    yoe = extract_required_yoe(text)
    return text, must_haves, nice_to_haves, yoe

# --- ISOLATED FRAGMENT FOR SKILL & REQUIREMENT CONTROLS ---
@st.fragment
def render_requirements_editor(uploaded_jd):
    if not uploaded_jd:
        return None, [], [], 0.0, False

    st.markdown("---")
    st.subheader("Auto-Extracted Job Requirements")
    st.caption("Review and adjust the extracted requirements before analyzing candidates.")
    
    jd_bytes = uploaded_jd.getvalue()
    jd_text, auto_skills, auto_nice_to_haves, auto_yoe = cached_parse_jd(jd_bytes, uploaded_jd.name)

    if "custom_skills" not in st.session_state:
        st.session_state.custom_skills = []
    if "selected_skills" not in st.session_state:
        st.session_state.selected_skills = auto_skills.copy()
    if "selected_nice_to_haves" not in st.session_state:
        st.session_state.selected_nice_to_haves = auto_nice_to_haves.copy()
    if "current_jd_name" not in st.session_state:
        st.session_state.current_jd_name = uploaded_jd.name

    if st.session_state.current_jd_name != uploaded_jd.name:
        st.session_state.current_jd_name = uploaded_jd.name
        st.session_state.custom_skills = [] 
        st.session_state.selected_skills = auto_skills.copy() 
        st.session_state.selected_nice_to_haves = auto_nice_to_haves.copy()

    col1, col2 = st.columns([3, 1])
    with col1:
        def add_custom_skill():
            new_s = st.session_state.new_skill_input.strip()
            if new_s:
                if new_s not in st.session_state.custom_skills:
                    st.session_state.custom_skills.append(new_s)
                if new_s not in st.session_state.selected_skills:
                    st.session_state.selected_skills.append(new_s)
            st.session_state.new_skill_input = ""

        def sync_skills():
            st.session_state.selected_skills = st.session_state.skills_widget

        st.text_input(
            "Type a missing skill (e.g., 'Next.js') and press Enter:", 
            key="new_skill_input", 
            on_change=add_custom_skill
        )
        
        all_options = list(set(auto_skills + st.session_state.custom_skills + ["Python", "AWS", "SQL", "Docker", "Kubernetes", "React", "Java", "Next.js"]))
        safe_defaults = [s for s in st.session_state.selected_skills if s in all_options]

        must_have_skills = st.multiselect(
            "Must-Have Skills (Click 'X' to remove):",
            options=all_options,
            default=safe_defaults,
            key="skills_widget",
            on_change=sync_skills
        )

        # Nice-to-have multiselect
        nice_to_have_skills = st.multiselect(
            "Nice-to-Have Skills (bonus, no penalty):",
            options=list(set(auto_nice_to_haves + all_options)),
            default=[s for s in st.session_state.selected_nice_to_haves if s in list(set(auto_nice_to_haves + all_options))],
            key="nice_to_have_widget"
        )
        
        strict_mode = st.checkbox("Strict Mode: Completely hide candidates missing ANY of these skills.")
        
    with col2:
        target_yoe = st.number_input("Required Years of Experience:", min_value=0.0, value=float(auto_yoe), step=0.5, format="%.1f")

    return jd_text, must_have_skills, nice_to_have_skills, target_yoe, strict_mode


# --- HELPER: Extract contact info from sections ---
def _extract_contact_info(sections_dict: dict) -> dict:
    full_search_text = " ".join([str(v) for v in sections_dict.values() if v])
    
    email_match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', full_search_text)
    phone_match = re.search(r'(?:\+\d{1,3}\s?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}|\+?\d{10,13}', full_search_text)
    linkedin_match = re.search(r'(?:https?:\/\/)?(?:www\.)?linkedin\.com\/in\/[\w\-_%]+', full_search_text)

    return {
        "email": email_match.group(0) if email_match else "N/A",
        "phone": phone_match.group(0) if phone_match else "N/A",
        "linkedin": linkedin_match.group(0) if linkedin_match else "N/A"
    }


# --- VISUAL DOSSIER: Render Candidate Card ---
def render_candidate_dossier(res: dict, must_have_skills: list):
    """Render a single candidate's full visual inspection dossier."""
    audit = res.get("audit", {})
    contact = _extract_contact_info(res.get("sections", {}))
    
    # --- HEADER: Name + Score + Contact ---
    candidate_name = res["file_name"].replace(".pdf", "").replace(".docx", "").replace("_", " ").title()
    
    st.markdown(f"#### 📋 {candidate_name}")
    
    # Contact info row
    contact_parts = []
    if contact["email"] != "N/A":
        contact_parts.append(f"📧 {contact['email']}")
    if contact["phone"] != "N/A":
        contact_parts.append(f"📱 {contact['phone']}")
    if contact["linkedin"] != "N/A":
        contact_parts.append(f"🔗 {contact['linkedin']}")
    if contact_parts:
        st.caption(" · ".join(contact_parts))
    
    # --- TOP METRICS ROW ---
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        score = res["final_score_pct"]
        delta_color = "normal" if score >= 65 else ("off" if score >= 50 else "inverse")
        st.metric("Final Score", f"{score}%", delta=f"{'🟢' if score >= 65 else '🟡' if score >= 50 else '🔴'}", delta_color="off")
    with m2:
        st.metric("Experience", f"{res['candidate_yoe']} yrs")
    with m3:
        st.metric("Skills Matched", f"{len(res['matched_skills'])}/{len(res['matched_skills']) + len(res['missing_skills'])}")
    with m4:
        bonus_count = len(res.get("nice_to_have_matched", []))
        st.metric("Bonus Skills", f"+{bonus_count}")

    # --- EXPLAINABLE SCORING (Raw Sub-Scores) ---
    st.markdown("##### 🧮 Explainable Sub-Scores")
    subscores = audit.get("subscores", {})
    s1, s2, s3, s4 = st.columns(4)
    with s1:
        st.metric("Skill Match", f"{subscores.get('skill_match', 0)} / 35 pts")
    with s2:
        st.metric("Recent Exp Match", f"{subscores.get('recent_exp', 0)} / 45 pts")
    with s3:
        st.metric("Older Exp Match", f"{subscores.get('older_exp', 0)} / 20 pts")
    with s4:
        st.metric("Keyword BM25", f"{subscores.get('bm25_keyword', 0)} / 100 pts")

    st.markdown("---")
    
    # --- SCORE BREAKDOWN WATERFALL (Horizontal Stacked Bar) ---
    st.markdown("##### 📊 Score Breakdown Waterfall")
    
    composite_base = audit.get("composite_base_pct", 0)
    penalty = audit.get("must_have_penalty_pct", 0)
    bonus = audit.get("nice_to_have_bonus_pct", 0)
    yoe_mod = audit.get("yoe_modifier_pct", 0)
    final = audit.get("calibrated_final_pct", res["final_score_pct"])
    
    # Build waterfall visualization using HTML/CSS
    bar_max = max(composite_base, 100)
    
    # Positive base
    base_width = max(0, (composite_base / bar_max) * 100) if bar_max > 0 else 0
    # Penalty is negative
    penalty_abs = abs(penalty)
    penalty_width = (penalty_abs / bar_max) * 100 if bar_max > 0 else 0
    # Bonus is positive
    bonus_width = (abs(bonus) / bar_max) * 100 if bar_max > 0 else 0
    # YOE modifier can be positive or negative
    yoe_abs = abs(yoe_mod)
    yoe_width = (yoe_abs / bar_max) * 100 if bar_max > 0 else 0
    yoe_color = "#4CAF50" if yoe_mod >= 0 else "#FF5722"
    
    waterfall_html = f"""
    <div style="margin: 10px 0 20px 0;">
        <div style="display: flex; align-items: center; margin-bottom: 8px;">
            <span style="width: 140px; font-size: 13px; color: #aaa;">Base Score</span>
            <div style="flex: 1; background: #1a1a2e; border-radius: 6px; height: 28px; position: relative; overflow: hidden;">
                <div style="width: {base_width:.1f}%; height: 100%; background: linear-gradient(90deg, #667eea, #764ba2); border-radius: 6px; display: flex; align-items: center; justify-content: flex-end; padding-right: 8px;">
                    <span style="color: white; font-size: 12px; font-weight: 600;">{composite_base:.1f}%</span>
                </div>
            </div>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 8px;">
            <span style="width: 140px; font-size: 13px; color: #aaa;">Skill Penalty</span>
            <div style="flex: 1; background: #1a1a2e; border-radius: 6px; height: 28px; position: relative; overflow: hidden;">
                <div style="width: {penalty_width:.1f}%; height: 100%; background: linear-gradient(90deg, #FF5722, #f44336); border-radius: 6px; display: flex; align-items: center; justify-content: flex-end; padding-right: 8px;">
                    <span style="color: white; font-size: 12px; font-weight: 600;">{penalty:.1f}%</span>
                </div>
            </div>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 8px;">
            <span style="width: 140px; font-size: 13px; color: #aaa;">Bonus Skills</span>
            <div style="flex: 1; background: #1a1a2e; border-radius: 6px; height: 28px; position: relative; overflow: hidden;">
                <div style="width: {bonus_width:.1f}%; height: 100%; background: linear-gradient(90deg, #00C853, #69F0AE); border-radius: 6px; display: flex; align-items: center; justify-content: flex-end; padding-right: 8px;">
                    <span style="color: white; font-size: 12px; font-weight: 600;">+{bonus:.1f}%</span>
                </div>
            </div>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 8px;">
            <span style="width: 140px; font-size: 13px; color: #aaa;">YOE Modifier</span>
            <div style="flex: 1; background: #1a1a2e; border-radius: 6px; height: 28px; position: relative; overflow: hidden;">
                <div style="width: {yoe_width:.1f}%; height: 100%; background: linear-gradient(90deg, {yoe_color}, {yoe_color}99); border-radius: 6px; display: flex; align-items: center; justify-content: flex-end; padding-right: 8px;">
                    <span style="color: white; font-size: 12px; font-weight: 600;">{'+' if yoe_mod >= 0 else ''}{yoe_mod:.1f}%</span>
                </div>
            </div>
        </div>
        <div style="display: flex; align-items: center; border-top: 1px solid #333; padding-top: 8px;">
            <span style="width: 140px; font-size: 14px; color: #fff; font-weight: 700;">FINAL</span>
            <div style="flex: 1; background: #1a1a2e; border-radius: 6px; height: 32px; position: relative; overflow: hidden;">
                <div style="width: {min(100, (final / bar_max) * 100):.1f}%; height: 100%; background: linear-gradient(90deg, #FFD700, #FFA000); border-radius: 6px; display: flex; align-items: center; justify-content: flex-end; padding-right: 8px;">
                    <span style="color: #1a1a2e; font-size: 13px; font-weight: 700;">{final:.1f}%</span>
                </div>
            </div>
        </div>
    </div>
    """
    st.markdown(waterfall_html, unsafe_allow_html=True)

    # --- SECTION SIMILARITY BAR CHART ---
    st.markdown("##### 🎯 Section Match Analysis")
    
    skills_sim = audit.get("skills_similarity_pct", 0)
    recent_sim = audit.get("recent_exp_similarity_pct", 0)
    older_sim = audit.get("older_exp_similarity_pct", 0)
    raw_bm25 = audit.get("raw_bm25_pct", 0)

    section_html = f"""
    <div style="margin: 10px 0 20px 0;">
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <span style="width: 160px; font-size: 12px; color: #aaa;">Skills Section</span>
            <div style="flex: 1; background: #1a1a2e; border-radius: 4px; height: 22px; overflow: hidden;">
                <div style="width: {min(100, skills_sim):.1f}%; height: 100%; background: #42A5F5; border-radius: 4px; display: flex; align-items: center; padding-left: 6px;">
                    <span style="color: white; font-size: 11px;">{skills_sim:.1f}%</span>
                </div>
            </div>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <span style="width: 160px; font-size: 12px; color: #aaa;">Recent Experience</span>
            <div style="flex: 1; background: #1a1a2e; border-radius: 4px; height: 22px; overflow: hidden;">
                <div style="width: {min(100, recent_sim):.1f}%; height: 100%; background: #66BB6A; border-radius: 4px; display: flex; align-items: center; padding-left: 6px;">
                    <span style="color: white; font-size: 11px;">{recent_sim:.1f}%</span>
                </div>
            </div>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <span style="width: 160px; font-size: 12px; color: #aaa;">Older Experience</span>
            <div style="flex: 1; background: #1a1a2e; border-radius: 4px; height: 22px; overflow: hidden;">
                <div style="width: {min(100, older_sim):.1f}%; height: 100%; background: #AB47BC; border-radius: 4px; display: flex; align-items: center; padding-left: 6px;">
                    <span style="color: white; font-size: 11px;">{older_sim:.1f}%</span>
                </div>
            </div>
        </div>
        <div style="display: flex; align-items: center; margin-bottom: 6px;">
            <span style="width: 160px; font-size: 12px; color: #aaa;">BM25 Keywords</span>
            <div style="flex: 1; background: #1a1a2e; border-radius: 4px; height: 22px; overflow: hidden;">
                <div style="width: {min(100, raw_bm25):.1f}%; height: 100%; background: #FFA726; border-radius: 4px; display: flex; align-items: center; padding-left: 6px;">
                    <span style="color: white; font-size: 11px;">{raw_bm25:.1f}%</span>
                </div>
            </div>
        </div>
    </div>
    """
    st.markdown(section_html, unsafe_allow_html=True)

    # --- SKILL BADGE MATRIX ---
    st.markdown("##### 🏷️ Skill Badge Matrix")
    
    verified = res.get("contextual_skills", [])
    stuffed = res.get("stuffed_skills", [])
    missing = res.get("missing_skills", [])
    bonus_skills = res.get("nice_to_have_matched", [])

    badges_html = '<div style="display: flex; flex-wrap: wrap; gap: 8px; margin: 10px 0 20px 0;">'
    
    for skill in verified:
        badges_html += f'''
        <div style="display: inline-flex; align-items: center; gap: 5px; padding: 5px 12px; 
            background: linear-gradient(135deg, #1b5e20, #2e7d32); border-radius: 20px; 
            border: 1px solid #4CAF50; font-size: 12px; color: #C8E6C9;">
            <span style="font-size: 14px;">🟢</span> {skill} (Verified)
        </div>'''
        
    for skill in stuffed:
        badges_html += f'''
        <div style="display: inline-flex; align-items: center; gap: 5px; padding: 5px 12px; 
            background: linear-gradient(135deg, #F57F17, #F9A825); border-radius: 20px; 
            border: 1px solid #FFCA28; font-size: 12px; color: #FFFDE7;">
            <span style="font-size: 14px;">🟡</span> {skill} (Stuffed)
        </div>'''
    
    for skill in missing:
        badges_html += f'''
        <div style="display: inline-flex; align-items: center; gap: 5px; padding: 5px 12px; 
            background: linear-gradient(135deg, #b71c1c, #c62828); border-radius: 20px; 
            border: 1px solid #f44336; font-size: 12px; color: #FFCDD2;">
            <span style="font-size: 14px;">❌</span> {skill}
        </div>'''

    for skill in bonus_skills:
        badges_html += f'''
        <div style="display: inline-flex; align-items: center; gap: 5px; padding: 5px 12px; 
            background: linear-gradient(135deg, #01579B, #0277BD); border-radius: 20px; 
            border: 1px solid #29B6F6; font-size: 12px; color: #B3E5FC;">
            <span style="font-size: 14px;">⭐</span> {skill}
        </div>'''

    badges_html += '</div>'
    
    # Legend
    badges_html += '''
    <div style="display: flex; gap: 16px; font-size: 11px; color: #888; margin-bottom: 10px;">
        <span>🟢 Verified (In Context)</span> <span>🟡 Stuffed (Listed Only)</span> <span>❌ Missing</span> <span>⭐ Bonus</span>
    </div>'''
    
    st.markdown(badges_html, unsafe_allow_html=True)


def main():
    st.title("AI-Powered ATS Resume Matcher & Scorer")
    st.markdown("Upload a Job Description to auto-extract requirements, then score candidate resumes using Hybrid Dense/Sparse AI matching.")

    # SIDEBAR: UPLOADS & WEIGHTS
    st.sidebar.header("Configuration & Uploads")
    uploaded_jd = st.sidebar.file_uploader("Upload Job Description (.pdf, .docx, .txt)", type=["pdf", "docx", "txt"])
    uploaded_cvs = st.sidebar.file_uploader("Upload Candidate CVs (.pdf, .docx)", type=["pdf", "docx"], accept_multiple_files=True)

    st.sidebar.markdown("---")
    st.sidebar.subheader("Scoring Controls")
    vector_weight = st.sidebar.slider("Semantic Vector Weight", 0.0, 1.0, 0.6)
    bm25_weight = st.sidebar.slider("Keyword BM25 Weight", 0.0, 1.0, 0.4)

    st.sidebar.markdown("---")
    st.sidebar.subheader("Penalty Tuning")
    penalty_severity = st.sidebar.slider(
        "Must-Have Penalty Severity",
        min_value=0.0, max_value=0.50, value=0.15, step=0.05,
        help="Maximum % deducted for missing ALL must-have skills. 0 = no penalty, 0.50 = harsh penalty."
    )

    jd_text, must_have_skills, nice_to_have_skills, target_yoe, strict_mode = render_requirements_editor(uploaded_jd)

    # SCORING EXECUTION
    if uploaded_jd and uploaded_cvs:
        st.markdown("---")
        if st.button("Run ATS Analysis", type="primary", use_container_width=True):
            
            with st.status("Processing documents in memory and calculating hybrid scores...", expanded=True) as status:
                registry = get_model_registry()
                parser = registry["parser"]
                scorer = registry["scorer"]
                
                candidates = []
                for cv_file in uploaded_cvs:
                    cv_bytes = cv_file.getvalue()
                    parsed_data = parser.parse_cv(cv_bytes, file_name=cv_file.name)
                    if parsed_data:
                        candidates.append(parsed_data)

                if not candidates:
                    status.update(label="Extraction Failed", state="error")
                    st.error("Could not extract text from the uploaded CVs.")
                    return

                # Consolidated scoring with full audit trail
                final_results = scorer.score_candidates(
                    jd_text=jd_text,
                    candidates=candidates,
                    must_have_skills=must_have_skills,
                    nice_to_have_skills=nice_to_have_skills,
                    target_yoe=target_yoe,
                    vector_weight=vector_weight,
                    bm25_weight=bm25_weight,
                    penalty_severity=penalty_severity
                )

                # Apply strict mode filter (display-only, not a score modification)
                if strict_mode:
                    final_results = [r for r in final_results if not r.get("missing_skills")]
                
                st.session_state.final_results = final_results
                st.session_state.must_have_skills = must_have_skills
                status.update(label=f"Successfully analyzed {len(final_results)} candidate profiles.", state="complete", expanded=False)

        # DISPLAY RESULTS IF AVAILABLE IN SESSION STATE
        if "final_results" in st.session_state and st.session_state.final_results:
            final_results = st.session_state.final_results
            
            st.markdown("### Candidate Ranking Results")

            display_data = []
            for idx, res in enumerate(final_results, start=1):
                ctx_len = len(res.get("contextual_skills", []))
                stuffed_len = len(res.get("stuffed_skills", []))
                
                contextual_str = f"{ctx_len} Verified"
                if stuffed_len > 0:
                    contextual_str += f", {stuffed_len} Stuffed"

                missing_str = ", ".join(res["missing_skills"][:3]) if res["missing_skills"] else "None"
                if len(res["missing_skills"]) > 3:
                    missing_str += f" (+{len(res['missing_skills']) - 3} more)"

                display_data.append({
                    "Rank": idx,
                    "Candidate File": res["file_name"],
                    "Final Match %": res['final_score_pct'],
                    "Est. YOE": f"{res['candidate_yoe']} yrs",
                    "Skills Validated": contextual_str,
                    "Missing Skills": missing_str
                })

            df_results = pd.DataFrame(display_data)
            st.dataframe(df_results, use_container_width=True, hide_index=True)

            # --- PROFESSIONAL EXCEL REPORT EXPORT ---
            export_rows = []
            for idx, res in enumerate(final_results, start=1):
                score = round(res.get("final_score_pct", 0.0), 2)
                contact = _extract_contact_info(res.get("sections", {}))
                audit = res.get("audit", {})
                subscores = audit.get("subscores", {})
                
                contextual_skills = ", ".join(res.get("contextual_skills", [])) if res.get("contextual_skills") else "None"
                stuffed_skills = ", ".join(res.get("stuffed_skills", [])) if res.get("stuffed_skills") else "None"
                missing_skills_full = ", ".join(res.get("missing_skills", [])) if res.get("missing_skills") else "None specified"
                nice_matched_full = ", ".join(res.get("nice_to_have_matched", [])) if res.get("nice_to_have_matched") else "None"
                
                export_rows.append({
                    "Rank": idx,
                    "Candidate File": res.get("file_name", "Unknown").replace(".pdf", "").replace(".docx", "").replace("_", " ").title(),
                    "Email Address": contact["email"],
                    "Phone Number": contact["phone"],
                    "LinkedIn Profile": contact["linkedin"],
                    "Final Match Score (%)": score,
                    "Estimated Experience (Years)": res.get("candidate_yoe", 0.0),
                    "Verified Skills": contextual_skills,
                    "Stuffed Skills": stuffed_skills,
                    "Missing Skills": missing_skills_full,
                    "Bonus Skills Matched": nice_matched_full,
                    "Skill Match (X/35 pts)": subscores.get("skill_match", ""),
                    "Recency (X/45 pts)": subscores.get("recent_exp", ""),
                    "Experience (X/20 pts)": subscores.get("older_exp", ""),
                    "Keyword BM25 (X/100 pts)": subscores.get("bm25_keyword", ""),
                    "Base Score (%)": audit.get("composite_base_pct", ""),
                    "Skill Penalty (%)": audit.get("must_have_penalty_pct", ""),
                    "Bonus (%)": audit.get("nice_to_have_bonus_pct", ""),
                    "YOE Modifier (%)": audit.get("yoe_modifier_pct", "")
                })

            df_export = pd.DataFrame(export_rows)

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_export.to_excel(writer, index=False, sheet_name='Candidate Rankings')
            
            output.seek(0)
            
            wb = openpyxl.load_workbook(output)
            ws = wb.active

            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Soft red for < 50%
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') # Soft green for >= 65%

            for row_idx in range(2, ws.max_row + 1):
                score_cell = ws.cell(row=row_idx, column=6)
                try:
                    score_val = float(score_cell.value)
                    if score_val < 50.0:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = red_fill
                    elif score_val >= 65.0:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = green_fill
                except (ValueError, TypeError):
                    pass

            # AUTO-FIT COLUMN WIDTHS
            for col in ws.columns:
                max_length = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max(max_length + 4, 15)

            final_excel_output = io.BytesIO()
            wb.save(final_excel_output)
            final_excel_output.seek(0)

            st.download_button(
                label="📥 Download Styled Excel Recruiter Report (.xlsx)",
                data=final_excel_output,
                file_name="executive_ats_recruitment_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )

            # --- VISUAL CANDIDATE DOSSIER ---
            st.markdown("---")
            st.markdown("### 🔍 Candidate Intelligence Dossier")
            st.caption("Select a candidate below to view their complete scoring breakdown with visual analytics.")

            candidate_options = [res["file_name"] for res in final_results]
            selected_candidate_name = st.selectbox("Choose Candidate File:", options=candidate_options, key="dossier_select")

            selected_res = next((res for res in final_results if res["file_name"] == selected_candidate_name), None)

            if selected_res:
                render_candidate_dossier(selected_res, must_have_skills)

if __name__ == "__main__":
    main()