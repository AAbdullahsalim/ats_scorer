import os
import io
import re
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import streamlit as st

from sentence_transformers import SentenceTransformer
from keybert import KeyBERT

# Core Engine Imports
from src.parser import ResumeParser, extract_must_haves_with_keybert, extract_required_yoe
from src.scorer import HybridScorer, evaluate_must_haves, apply_must_have_penalty, estimate_candidate_yoe, apply_yoe_modifier

# 1. Page Configuration
st.set_page_config(
    page_title="AI ATS Resume Matcher", 
    page_icon="⚡", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- CENTRALIZED LAZY-LOADED MODEL REGISTRY ---
@st.cache_resource(show_spinner="Initializing AI Model Registry...")
def get_model_registry():
    transformer = SentenceTransformer("all-MiniLM-L6-v2")
    keybert_model = KeyBERT(model=transformer)
    scorer = HybridScorer(vector_model=transformer)
    parser = ResumeParser()
    return {
        "transformer": transformer,
        "keybert": keybert_model,
        "scorer": scorer,
        "parser": parser
    }

# --- TOP-LEVEL CACHED PARSING FUNCTION ---
@st.cache_data
def cached_parse_jd(jd_bytes: bytes, file_name: str):
    registry = get_model_registry()
    parser = registry["parser"]
    keybert_model = registry["keybert"]
    
    text = parser.parse_jd(jd_bytes, file_name=file_name)
    skills = extract_must_haves_with_keybert(text, keybert_model=keybert_model)
    yoe = extract_required_yoe(text)
    return text, skills, yoe

# --- ISOLATED FRAGMENT FOR SKILL & REQUIREMENT CONTROLS ---
@st.fragment
def render_requirements_editor(uploaded_jd):
    if not uploaded_jd:
        return None, [], 0.0, False

    st.markdown("---")
    st.subheader("Auto-Extracted Job Requirements")
    st.caption("Review and adjust the extracted requirements before analyzing candidates.")
    
    jd_bytes = uploaded_jd.getvalue()
    jd_text, auto_skills, auto_yoe = cached_parse_jd(jd_bytes, uploaded_jd.name)

    if "custom_skills" not in st.session_state:
        st.session_state.custom_skills = []
    if "selected_skills" not in st.session_state:
        st.session_state.selected_skills = auto_skills.copy()
    if "current_jd_name" not in st.session_state:
        st.session_state.current_jd_name = uploaded_jd.name

    if st.session_state.current_jd_name != uploaded_jd.name:
        st.session_state.current_jd_name = uploaded_jd.name
        st.session_state.custom_skills = [] 
        st.session_state.selected_skills = auto_skills.copy() 

    col1, col2 = st.columns([3, 1])
    with col1:
        def add_custom_skill():
            new_s = st.session_state.new_skill_input.strip()
            if new_s:
                if new_s not in st.session_state.custom_skills:
                    st.session_state.custom_skills.append(new_s)
                if new_s not in st.session_state.selected_skills:
                    st.session_state.selected_skills.append(new_s)
            st.session_state.new_skill_input = ""

        def sync_skills():
            st.session_state.selected_skills = st.session_state.skills_widget

        st.text_input(
            "Type a missing skill (e.g., 'Next.js') and press Enter:", 
            key="new_skill_input", 
            on_change=add_custom_skill
        )
        
        all_options = list(set(auto_skills + st.session_state.custom_skills + ["Python", "AWS", "SQL", "Docker", "Kubernetes", "React", "Java", "Next.js"]))
        safe_defaults = [s for s in st.session_state.selected_skills if s in all_options]

        must_have_skills = st.multiselect(
            "Must-Have Skills (Click 'X' to remove):",
            options=all_options,
            default=safe_defaults,
            key="skills_widget",
            on_change=sync_skills
        )
        
        strict_mode = st.checkbox("Strict Mode: Completely hide candidates missing ANY of these skills.")
        
    with col2:
        target_yoe = st.number_input("Required Years of Experience:", min_value=0.0, value=float(auto_yoe), step=0.5, format="%.1f")

    return jd_text, must_have_skills, target_yoe, strict_mode


def main():
    st.title("AI-Powered ATS Resume Matcher & Scorer")
    st.markdown("Upload a Job Description to auto-extract requirements, then score candidate resumes using Hybrid Dense/Sparse AI matching.")

    # SIDEBAR: UPLOADS & WEIGHTS
    st.sidebar.header("Configuration & Uploads")
    uploaded_jd = st.sidebar.file_uploader("Upload Job Description (.pdf, .docx, .txt)", type=["pdf", "docx", "txt"])
    uploaded_cvs = st.sidebar.file_uploader("Upload Candidate CVs (.pdf, .docx)", type=["pdf", "docx"], accept_multiple_files=True)

    st.sidebar.markdown("---")
    st.sidebar.subheader("Base Scoring Weights")
    vector_weight = st.sidebar.slider("Semantic Vector Weight", 0.0, 1.0, 0.6)
    bm25_weight = st.sidebar.slider("Keyword BM25 Weight", 0.0, 1.0, 0.4)

    jd_text, must_have_skills, target_yoe, strict_mode = render_requirements_editor(uploaded_jd)

    # SCORING EXECUTION
    if uploaded_jd and uploaded_cvs:
        st.markdown("---")
        if st.button("Run ATS Analysis", type="primary", use_container_width=True):
            
            with st.status("Processing documents in memory and calculating hybrid scores...", expanded=True) as status:
                registry = get_model_registry()
                parser = registry["parser"]
                scorer = registry["scorer"]
                
                candidates = []
                for cv_file in uploaded_cvs:
                    cv_bytes = cv_file.getvalue()
                    parsed_data = parser.parse_cv(cv_bytes, file_name=cv_file.name)
                    if parsed_data:
                        candidates.append(parsed_data)

                if not candidates:
                    status.update(label="Extraction Failed", state="error")
                    st.error("Could not extract text from the uploaded CVs.")
                    return

                base_results = scorer.score_candidates(
                    jd_text=jd_text,
                    candidates=candidates,
                    vector_weight=vector_weight,
                    bm25_weight=bm25_weight
                )

                final_results = []
                for res in base_results:
                    cv_full_text = " ".join(res.get("sections", {}).values())
                    base_score = res['final_score_pct']
                    
                    must_have_eval = evaluate_must_haves(cv_full_text, must_have_skills)
                    if strict_mode and must_have_eval["ratio"] < 1.0:
                        continue
                        
                    score_after_skills = apply_must_have_penalty(base_score, must_have_eval["ratio"])
                    candidate_yoe = estimate_candidate_yoe(cv_full_text)
                    final_score = apply_yoe_modifier(score_after_skills, candidate_yoe, target_yoe)
                    
                    res["base_score"] = base_score
                    res["final_score_pct"] = final_score
                    res["matched_skills"] = must_have_eval["matched"]
                    res["missing_skills"] = must_have_eval["missing"]
                    res["candidate_yoe"] = candidate_yoe
                    final_results.append(res)

                final_results = sorted(final_results, key=lambda x: x["final_score_pct"], reverse=True)
                
                st.session_state.final_results = final_results
                st.session_state.must_have_skills = must_have_skills
                status.update(label=f"Successfully analyzed {len(final_results)} candidate profiles.", state="complete", expanded=False)

        # DISPLAY RESULTS IF AVAILABLE IN SESSION STATE
        if "final_results" in st.session_state and st.session_state.final_results:
            final_results = st.session_state.final_results
            
            st.markdown("### Candidate Ranking Results")

            display_data = []
            for idx, res in enumerate(final_results, start=1):
                matched_str = ", ".join(res["matched_skills"][:3]) if res["matched_skills"] else "None"
                if len(res["matched_skills"]) > 3:
                    matched_str += f" (+{len(res['matched_skills']) - 3} more)"

                missing_str = ", ".join(res["missing_skills"][:3]) if res["missing_skills"] else "None"
                if len(res["missing_skills"]) > 3:
                    missing_str += f" (+{len(res['missing_skills']) - 3} more)"

                display_data.append({
                    "Rank": idx,
                    "Candidate File": res["file_name"],
                    "Final Match %": res['final_score_pct'],
                    "Est. YOE": f"{res['candidate_yoe']} yrs",
                    "Matched Skills": matched_str,
                    "Missing Skills": missing_str
                })

            df_results = pd.DataFrame(display_data)
            st.dataframe(df_results, use_container_width=True, hide_index=True)

            # --- PROFESSIONAL EXCEL REPORT EXPORT (WITH AUTO-FIT, CONTACTS & ROW COLORING) ---
            export_rows = []
            for idx, res in enumerate(final_results, start=1):
                score = round(res.get("final_score_pct", 0.0), 2)
                
                sections_dict = res.get("sections", {})
                full_search_text = " ".join([str(v) for v in sections_dict.values() if v])
                
                email_match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', full_search_text)
                phone_match = re.search(r'(?:\+\d{1,3}\s?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}|\+?\d{10,13}', full_search_text)
                linkedin_match = re.search(r'(?:https?:\/\/)?(?:www\.)?linkedin\.com\/in\/[\w\-_%]+', full_search_text)

                email = email_match.group(0) if email_match else "N/A"
                phone = phone_match.group(0) if phone_match else "N/A"
                linkedin = linkedin_match.group(0) if linkedin_match else "N/A"
                
                matched_skills_full = ", ".join(res.get("matched_skills", [])) if res.get("matched_skills") else "None specified"
                missing_skills_full = ", ".join(res.get("missing_skills", [])) if res.get("missing_skills") else "None specified"
                
                export_rows.append({
                    "Rank": idx,
                    "Candidate File": res.get("file_name", "Unknown").replace(".pdf", "").replace(".docx", "").replace("_", " ").title(),
                    "Email Address": email,
                    "Phone Number": phone,
                    "LinkedIn Profile": linkedin,
                    "Final Match Score (%)": score,
                    "Estimated Experience (Years)": res.get("candidate_yoe", 0.0),
                    "Matched Skills": matched_skills_full,
                    "Missing Skills": missing_skills_full
                })

            df_export = pd.DataFrame(export_rows)

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_export.to_excel(writer, index=False, sheet_name='Candidate Rankings')
            
            output.seek(0)
            
            wb = openpyxl.load_workbook(output)
            ws = wb.active

            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Soft red for < 50%
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') # Soft green for >= 65%

            for row_idx in range(2, ws.max_row + 1):
                score_cell = ws.cell(row=row_idx, column=6)
                try:
                    score_val = float(score_cell.value)
                    if score_val < 50.0:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = red_fill
                    elif score_val >= 65.0:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = green_fill
                except (ValueError, TypeError):
                    pass

            # AUTO-FIT COLUMN WIDTHS
            for col in ws.columns:
                max_length = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max(max_length + 4, 15)

            final_excel_output = io.BytesIO()
            wb.save(final_excel_output)
            final_excel_output.seek(0)

            st.download_button(
                label="📥 Download Styled Excel Recruiter Report (.xlsx)",
                data=final_excel_output,
                file_name="executive_ats_recruitment_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )

            # DETAILED INSPECTION CARD SELECTOR
            st.markdown("---")
            st.markdown("### Candidate Skill Inspection Dossier")
            st.caption("Select a candidate below to view their complete, un-truncated skill breakdown.")

            candidate_options = [res["file_name"] for res in final_results]
            selected_candidate_name = st.selectbox("Choose Candidate File:", options=candidate_options, key="dossier_select")

            selected_res = next((res for res in final_results if res["file_name"] == selected_candidate_name), None)

            if selected_res:
                st.markdown(f"#### Dossier for: **{selected_res['file_name']}**")
                
                col_i1, col_i2, col_i3 = st.columns(3)
                with col_i1:
                    st.metric("Final Match Score", f"{selected_res['final_score_pct']}%")
                with col_i2:
                    st.metric("Estimated Experience", f"{selected_res['candidate_yoe']} yrs")
                with col_i3:
                    st.metric("Total Match Ratio", f"{len(selected_res['matched_skills'])} matched")

                st.markdown("##### Complete Matched Skills List:")
                if selected_res["matched_skills"]:
                    st.success(", ".join(selected_res["matched_skills"]))
                else:
                    st.info("No target skills matched.")

                st.markdown("##### Complete Missing Skills List:")
                if selected_res["missing_skills"]:
                    st.error(", ".join(selected_res["missing_skills"]))
                else:
                    st.success("No missing skills! Candidate covers all requirements.")

if __name__ == "__main__":
    main()